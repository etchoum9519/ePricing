
from concurrent.futures import ThreadPoolExecutor
from flask import Flask, request, jsonify
from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

from decimal import Decimal
import pandas as pd
import os,re,time,calendar,sys

from zoneinfo import ZoneInfo

TWOPLACES = Decimal("0.001")

NUMBER_PATTERN = re.compile(
    r"^\d{1,3}(\.\d{3})+,\d+$"
)

app = Flask(__name__)
latest_run = {
    "filepath": None,
    "einheit": None,
    "energy_type": None
}

# =========================
# HILFSFUNKTIONEN
# =========================

FREQ_MAP = {
    "1h": "h",
    "1D": "D",
    "1M": "MS"
}

base = "/home"

print("BASE:", base)
print("ALLE:", os.listdir(os.path.join(base, "pfad1", "pfad2", "pfad3")))

def normalize_dst_hourly(df, energy_type, tz="Europe/Berlin"):
    df = df.copy()
    if energy_type != "gas":
        df["Timestamp"] = pd.to_datetime(df["Timestamp"], utc=True).dt.tz_convert(tz)

    if energy_type == "gas":
        df["Timestamp"] -= pd.Timedelta(hours=6)
    
    return df

    
def granularity_to_freq(granularity):
    return FREQ_MAP.get(granularity, "h")
    

def safe_factor(value):
    if value is None:
        return Decimal("1")

    try:
        v = str(value).replace(",", ".").replace("%", "").replace(" ", "")
        return Decimal(v) / Decimal("100")
    except:
        return Decimal("1")


def get_interval_hours(granularity, timestamp=None, energy_type="strom"):

    if granularity == "1h":
        return Decimal("1")


    if granularity == "1M":

        ts = pd.Timestamp(timestamp)

        days = calendar.monthrange(
            ts.year,
            ts.month
        )[1]

        return Decimal(str(days * 24))


    if granularity == "1D":

        ts = pd.Timestamp(timestamp)

        if energy_type == "gas":
            ts = ts - pd.Timedelta(hours=6)

        d = ts.date()

        last_march = max(
            week[calendar.SUNDAY]
            for week in calendar.monthcalendar(d.year, 3)
            if week[calendar.SUNDAY] != 0
        )

        last_october = max(
            week[calendar.SUNDAY]
            for week in calendar.monthcalendar(d.year, 10)
            if week[calendar.SUNDAY] != 0
        )


        if (
            (energy_type == "strom" and d.month == 3 and d.day == last_march)
            or
            (energy_type == "gas" and d.month == 3 and d.day == last_march-2)
        ):
            return Decimal("23")


        if (
            (energy_type == "strom" and d.month == 10 and d.day == last_october)
            or
            (energy_type == "gas" and d.month == 10 and d.day == last_october-2)
        ):
            return Decimal("25")


        return Decimal("24")


    raise ValueError(
        f"Unbekannte Granularität: {granularity}"
    )


def convert_to_mw(values, source_unit, granularity="1h", energy_type = "strom", timestamps=None):

    source_unit = (source_unit or "").strip()

    power_factors = {
        "W": Decimal("0.000001"),
        "kW": Decimal("0.001"),
        "MW": Decimal("1"),
        "GW": Decimal("1000")
    }

    energy_factors = {
        "Wh": Decimal("0.000001"),
        "kWh": Decimal("0.001"),
        "MWh": Decimal("1"),
        "GWh": Decimal("1000")
    }

    result = []

    for i, v in enumerate(values):

        v = Decimal(str(v))

        if source_unit in power_factors:
            result.append(v * power_factors[source_unit])
            continue

        if source_unit not in energy_factors:
            raise ValueError(f"Unbekannte Einheit: {source_unit}")
        
        mwh = v * energy_factors[source_unit]
        
        
        ts = None if timestamps is None else timestamps[i]
        hours = get_interval_hours(
            granularity,
            ts,
            energy_type=energy_type
        )
        
        result.append(mwh / hours)

    return result


def round_for_sum(values, unit):

    if unit in ["MW", "MWh"]:
        step = Decimal("0.001")
    elif unit in ["kW", "kWh"]:
        step = Decimal("1")
    else:
        step = Decimal("0.001")

    return [
        Decimal(v).quantize(step)
        for v in values
    ]

    
def is_energy_unit(unit):

    return unit in [
        "Wh",
        "kWh",
        "MWh",
        "GWh"
    ]


def build_all_granularities(df, input_unit, source_granularity, energy_type):
    # =========================
    # FALL 1: 1H → bleibt feinste Basis
    # =========================
    if source_granularity == "1h":
        hourly_df = df.copy()
        daily_df = derive_daily(hourly_df, energy_type) # energy_type
        monthly_df = derive_monthly(hourly_df, energy_type, granularity = source_granularity)

    if source_granularity == "1D":
        daily_df = df.copy()
        hourly_df = normalize_to_hourly(daily_df, energy_type=energy_type, source_granularity="1D")
        print(hourly_df.iloc[0])
        print(hourly_df.iloc[-1])
        
        monthly_df = derive_monthly(hourly_df, energy_type, granularity = "1h")
        print(monthly_df)

    # =========================
    # FALL 3: 1M → KEINE feinere Granularität erzeugen
    # =========================
    elif source_granularity == "1M":
    
        monthly_df = df.copy()
        daily_df = derive_daily(monthly_df)
        hourly_df = normalize_to_hourly(monthly_df, energy_type=energy_type, source_granularity="1M")

    vollversorgung_daily_df = derive_daily(monthly_df)
    print(daily_df.loc[
        daily_df["Timestamp"].dt.strftime("%Y-%m-%d").isin([
            "2028-03-25",
            "2028-10-28"
        ])
    ])
    # =========================
    # CONDITIONAL GRANULARITY LOGIC
    # =========================
    return (
        hourly_df.reset_index(drop=True) if hourly_df is not None else None,
        daily_df.reset_index(drop=True),
        monthly_df.reset_index(drop=True),
        vollversorgung_daily_df.reset_index(drop=True)
        if vollversorgung_daily_df is not None else None
    )


def normalize_to_hourly(df, energy_type="strom", source_granularity=None):
    df["Timestamp"] = pd.to_datetime(df["Timestamp"])
    
    if df["Timestamp"].dt.tz is None:
        df["Timestamp"] = df["Timestamp"].dt.tz_localize(
            "Europe/Berlin",
            ambiguous="infer",
            nonexistent="shift_forward"
        )
    else:
        df["Timestamp"] = df["Timestamp"].dt.tz_convert("Europe/Berlin")

    if df["Timestamp"].dt.tz is None:
        df["Timestamp"] = df["Timestamp"].dt.tz_localize(
            "Europe/Berlin",
            ambiguous="infer",
            nonexistent="shift_forward"
        )

    df = df.sort_values("Timestamp")

    start_ts = df["Timestamp"].min()
    # FIX: +1 Tag Puffer damit Jahreswechsel (inkl. 01.01 05:00) enthalten bleibt
    last_ts = df["Timestamp"].max()

    start_timestamp = start_ts
    if energy_type == "gas":
        start_timestamp = (
            pd.Timestamp(start_ts)
            .tz_convert(None)
            .normalize()
            + pd.Timedelta(hours=6)
        ).tz_localize("Europe/Berlin")
        last_day = (
            pd.Timestamp(last_ts)
            .tz_localize(None)
            .to_period("M")
            .to_timestamp("M")
        )
    
        last_timestamp = (
            last_day
            + pd.Timedelta(days=1, hours=5)
        ).tz_localize("Europe/Berlin")
    
    else:  #strom

            start_timestamp = (
                pd.Timestamp(start_ts)
                .tz_localize(None)
                .to_period("D")
                .to_timestamp()
            ).tz_localize("Europe/Berlin")
            last_timestamp = (
                pd.Timestamp(last_ts)
                .tz_localize(None)
                .to_period("M")
                .to_timestamp("M")
                + pd.Timedelta(hours=23)
            ).tz_localize("Europe/Berlin")
    

    hourly_index = pd.date_range(
        start=start_timestamp,
        end=last_timestamp,
        freq="h",
        tz="Europe/Berlin"
    )

    df = (
        df.sort_values("Timestamp")
          .drop_duplicates(
              subset=["Timestamp"],
              keep="first"
          )
    )

    df = (
        df.set_index("Timestamp")
          .groupby(level=0)
          .first()
    )
    
    if energy_type == "gas":
        df.index = (
            df.index
            .tz_convert(None)
            .normalize()
            + pd.Timedelta(hours=6)
        ).tz_localize("Europe/Berlin")
    
    df = df.reindex(hourly_index, method="ffill")

    if "MW" in df.columns:
        df["MW"] = df["MW"].ffill().bfill()

    df.index.name = "Timestamp"

    return (
        df.reset_index()
          .sort_values("Timestamp")
          .reset_index(drop=True)
    )


def derive_daily(df, energy_type="strom"):
    df = df.copy()

    df["Timestamp"] = pd.to_datetime(df["Timestamp"])
    df["Timestamp"] = df["Timestamp"].dt.tz_localize(None)
    
    if (
        len(df) >= 2
        and (
            df["Timestamp"]
            .sort_values()
            .diff()
            .dt.days
            .median()
            >= 28
        )
    ):
        expanded = []

        for _, row in df.iterrows():

            month_start = (
                            pd.Timestamp(row["Timestamp"])
                            .tz_localize(None)
                            .to_period("M")
                            .to_timestamp()
                        )

            if energy_type == "gas":

                last_day = (
                    month_start
                    .to_period("M")
                    .to_timestamp("M")
                )

                last_timestamp = (
                    last_day
                    + pd.Timedelta(days=1, hours=5)
                )

            else:
                month_start = (
                    pd.Timestamp(row["Timestamp"])
                    .to_period("M")
                    .start_time
                )
                
                last_timestamp = (
                    month_start
                    + pd.offsets.MonthEnd(0)
                )

            days = pd.date_range(
                start=month_start,
                end=last_timestamp,
                freq="D"
            )

            expanded.append(
                pd.DataFrame({
                    "Timestamp": days,
                    "MW": row["MW"]
                })
            )

        return (
            pd.concat(expanded, ignore_index=True)
            .sort_values("Timestamp")
            .reset_index(drop=True)
        )

    return (
        df.groupby(df["Timestamp"].dt.floor("D"))
        .agg({"MW": "mean"})
        .reset_index()
    )


def derive_monthly(df, energy_type="strom", granularity="X"):
    df = df.copy()

    df["Timestamp"] = pd.to_datetime(df["Timestamp"])
    # GAS + 1h: vor Monatsbildung auf 00:00 normalisieren
    if energy_type == "gas" and granularity == "1h":
        df = df.copy()
        df["Timestamp"] = df["Timestamp"] - pd.Timedelta(hours=6)
        df["Timestamp"] = df["Timestamp"].dt.normalize()
    # FALL 1: Tageslastgang -> auf Monate aggregieren
    if (
        len(df) >= 2
        and (
            df["Timestamp"]
            .sort_values()
            .diff()
            .dt.days
            .median()
            <= 1
        )
    ):
        df["Month"] = df["Timestamp"].dt.to_period("M")

        monthly = (
            df.groupby("Month", as_index=False)
            .agg({"MW": "mean"})
        )

        monthly["Timestamp"] = monthly["Month"].dt.to_timestamp()

        return monthly[["Timestamp", "MW"]]
    # FALL 2: Monatslastgang -> nur propagieren
    
    return (
        df[["Timestamp", "MW"]]
        .reset_index(drop=True)
    )

    
def validate_hourly_series(timestamps):

    if len(timestamps) < 2:
        return None

    ts = pd.to_datetime(
        pd.Series(timestamps),
        format="%Y-%m-%d %H:%M:%S",
        errors="coerce"
    )

    ts = pd.Series(ts).dropna()
    
    # sauber lokalisieren (WICHTIG: fix statt infer!)
    
    diffs = (
        ts.diff()
        .dropna()
        .dt.total_seconds()
        .div(3600)
    )
    # DST EVENT DETECTION
    is_spring_forward = (
        diffs.between(1.98, 2.02)
    )
    
    is_fall_back = (
        diffs.abs() < 0.02
    )
    
    # normale Stundenwerte tolerieren
    is_normal_hour = (
        diffs.between(0.98, 1.02)
    )
    
    valid = (
        is_normal_hour
        | is_spring_forward
        | is_fall_back
    )
    
    invalid = diffs[~valid]

    if not invalid.empty:  

        idx = invalid.index[0]

        return {
            "ok": False,
            "prev": str(ts.iloc[idx - 1]),
            "curr": str(ts.iloc[idx]),
            "diff_hours": float(diffs.loc[idx])
        }

    return {
        "ok": True
    }


def parse_number(value):
    if pd.isna(value):
        return None

    s = str(value).strip()

    # 1) deutsches Format 1.234,56
    if NUMBER_PATTERN.match(s):
        s = s.replace(".", "").replace(",", ".")

    # 2) nur Komma = deutsch
    elif "," in s and "." not in s:
        s = s.replace(",", ".")

    # 3) US Format 1,234.56
    elif "," in s and "." in s:
        if s.rfind(",") < s.rfind("."):
            s = s.replace(",", "")
        else:
            s = s.replace(".", "").replace(",", ".")

    try:
        return Decimal(s) if s else None
    except:
        return None


MONTH_MAP = {
    "j": 1, "jan": 1, "januar": 1,
    "feb": 2, "februar": 2,
    "mär": 3, "maerz": 3, "märz": 3, "mar": 3, "mrz": 3,
    "apr": 4, "april": 4,
    "mai": 5,
    "jun": 6, "juni": 6,
    "jul": 7, "juli": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "okt": 10, "oktober": 10,
    "nov": 11, "november": 11,
    "dez": 12, "dezember": 12
}

def parse_month_year(value):

    original = str(value).strip()

    # Nicht registrierte Monatsformate blockieren
    # z.B. "Aug." oder "Sep."
    if re.search(r"[A-Za-zÄÖÜäöü]\.$", original):
        return pd.NaT

    s = original.lower()

    s = re.sub(r"[-_/]+", ".", s)
    s = re.sub(r"\s+", ".", s)
    s = re.sub(r"\.+", ".", s)
    s = s.strip(".")
    # Spezialfall: nur Monat (z.B. "Aug") → Jahr aus Auswahl ergänzen
    if re.fullmatch(r"[a-zäöü]+", s, re.IGNORECASE):
        selected_years = sorted(
            int(y) for y in ((request.get_json(silent=True) or {}).get("years") or [])
        )
    
        if selected_years:
            year = selected_years[0]
    
            # Bei mehrjährigen Lastgängen: nach Dezember beginnt automatisch das nächste Jahr
            if len(selected_years) > 1:
                month = MONTH_MAP.get(s)
                if month == 1:
                    year += 1
    
            s = f"{s}.{year}"
    #print(repr(s))
    m = re.match(
        r"^([a-zäöü]+|\d{1,2})[\.\-_/ ]+(\d{2,4})$",
        s,
        re.IGNORECASE
    )
    
    if not m:
        m = re.match(
            r"^(\d{2,4})[\.\-_/ ]*([a-zäöü]+|\d{1,2})$",
            s,
            re.IGNORECASE
        )
    
        if not m:
            return pd.NaT
    
        year_part, month_part = m.groups()
    else:
        month_part, year_part = m.groups()

    if month_part.isdigit():
        month = int(month_part)
    else:
        month = MONTH_MAP.get(month_part)

    if not month or not (1 <= month <= 12):
        return pd.NaT

    if not year_part.isdigit():
        return pd.NaT

    year = int(year_part)

    if year < 100:
        year += 2000

    return pd.Timestamp(year=year, month=month, day=1)

    
def parse_lastgang(raw_input, forward_factor, granularity):
    data = request.get_json()

    if not raw_input:
        return {"values": [], "timestamps": None}

    lines = [l.strip() for l in raw_input.splitlines() if l.strip()]

    if not lines:
        return {"values": [], "timestamps": None}

    first_parts = re.split(r"[;\t]+", lines[0])

    # =========================
    # 1 SPALTE (bleibt ok)
    # =========================
    if len(first_parts) == 1:
    
        values = []
    
        for line in lines:
    
            value = parse_number(line)
    
            if value is not None:
                values.append(
                    value * forward_factor
                )
    
        return {
            "values": values,
            "timestamps": None
        }
        

    # =========================
    # 2 SPALTEN
    # =========================
    
    SPLIT_PATTERN = re.compile(r"[;\t]+")
    
    rows = [
        [x.strip() for x in SPLIT_PATTERN.split(l)]
        for l in lines
    ]
    
    df = pd.DataFrame(rows)
    df = df.iloc[:, :2]
    df.columns = ["ts", "val"]
    
    df["val"] = df["val"].apply(parse_number)
    df["val"] = df["val"].fillna(Decimal("0"))
    
    df["val"] = df["val"].map(
        lambda x: x * forward_factor
    )
    
    df = df.dropna(subset=["ts"])
    
    first_ts = str(df["ts"].iloc[0]).strip()
    
    # =========================
    # MONATLICH
    # =========================
    
    # Nur prüfen, wenn tatsächlich ein Monatslastgang erkannt wurde
    monthly_candidate = False
    
    for ts_value in df["ts"]:
    
        ts_str = str(ts_value).strip()
    
        if not pd.isna(parse_month_year(ts_str)):
            monthly_candidate = True
            break
            
        
    if monthly_candidate:
    
        for row_index, ts_value in enumerate(df["ts"]):
    
            ts_str = str(ts_value).strip()
    
            parsed_check = parse_month_year(ts_str)
    
            if pd.isna(parsed_check):
    
                return {
                    "error": True,
                    "message":
                        f"Ungültiger Monatszeitstempel in Zeile {row_index + 1}: '{ts_str}'. "
                        f"Bitte selektieren Sie ein Jahr oder entfernen Sie unerlaubte Zeichen. "
                }
        if pd.isna(parsed_check):
            return {
                "error": True,
                "message":
                        f"Ungültiger Monatszeitstempel in Zeile {row_index + 1}: '{ts_str}'. "
                        f"Bitte selektieren Sie ein Jahr oder entfernen Sie unerlaubte Zeichen. "
            }
    
    if re.fullmatch(r"[a-zäöü]+", first_ts, re.IGNORECASE):
        parsed_month = pd.Timestamp.now()
    else:
        parsed_month = parse_month_year(first_ts)
    if not pd.isna(parsed_month):
    
        selected_years = sorted(
            int(y) for y in ((request.get_json(silent=True) or {}).get("years") or [])
        )
        
        if (
            selected_years
            and all(
                re.fullmatch(
                    r"[a-zäöü]+",
                    str(x).strip(),
                    re.IGNORECASE
                )
                for x in df["ts"]
            )
        ):
        
            current_year = selected_years[0]
            previous_month = None
        
            timestamps = []
        
            for m in df["ts"]:
        
                month = MONTH_MAP[str(m).strip().lower()]
        
                if previous_month is not None and month < previous_month:
                    current_year += 1
        
                timestamps.append(
                    pd.Timestamp(
                        year=current_year,
                        month=month,
                        day=1
                    )
                )
        
                previous_month = month
        
            df["ts"] = timestamps
        
        else:
        
            df["ts"] = df["ts"].apply(parse_month_year)

            #print(df["ts"])
            #print(df["ts"].apply(parse_month_year))
            # ungültige Monatszeitstempel entfernen
            # (Schutz gegen nicht erkannte Formate)
            df = df.dropna(subset=["ts"])
        
        df = df.dropna(subset=["ts"])
        
        df["ts"] = pd.to_datetime(df["ts"])
        
        df["ts"] = df["ts"].dt.to_period("M").dt.to_timestamp()
    
    # =========================
    # TÄGLICH
    # =========================
    
    elif granularity == "1D":
        df["ts"] = pd.to_datetime(
            df["ts"],
            dayfirst=True,
            errors="coerce"
        ).dt.normalize()
    
    # =========================
    # STÜNDLICH
    # =========================
    
    else:
    
        df["ts"] = pd.to_datetime(
            df["ts"],
            dayfirst=True,
            errors="coerce"
        )
    
        df["ts"] = df["ts"].dt.round("h")
    
    df["val"] = (
        pd.to_numeric(
            df["val"],
            errors="coerce"
        )
        .fillna(0)
    )
    
    df = df.reset_index(drop=True)
    
    return {
        "values": df["val"].tolist(),
        "timestamps": df["ts"].tolist()
    }


# =========================
# STARTZEIT JE NACH TYP
# =========================
def get_start_hour(energy_type):

    # Strom startet 00:00
    if energy_type == "strom":
        return "00:00"

    # Gas startet 06:00
    return "06:00"



def normalize_gas_day_window(df, energy_type):
    if energy_type != "gas":
        return df

    df = df.copy()

    # bringt wieder zurück in 06:00–05:00 Sicht (Frontend korrekt)
    df["Timestamp"] = df["Timestamp"] + pd.Timedelta(hours=6)

    return df
    

# =========================
# EXCEL EXPORT
# =========================
def save_excel_file(df, pricing_name, energy_type, granularity, export_dir):

    if df is None or df.empty:
        print(f"⚠️ Kein DataFrame für {granularity}")
        return

    export_df = df
    
    if "MW" in export_df.columns:
        export_df["MW"] = export_df["MW"].apply(
            lambda x: Decimal(str(x)).quantize(TWOPLACES)
        )

    if "Timestamp" in export_df.columns:
        try:
            export_df["Timestamp"] = (
                export_df["Timestamp"]
                .dt.tz_localize(None)
            )
        except Exception:
            pass

    safe_name = pricing_name or "lastgang_export"
    prefix = "Gas" if energy_type == "gas" else "Strom"

    # Ordner anlegen
    os.makedirs(export_dir, exist_ok=True)

    # Zeitstempel z.B. 20260531_1613
    timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M")

    filename = (
        f"{prefix}_{safe_name}_{timestamp}_{granularity}.xlsx"
    )

    filepath = os.path.join(
        export_dir,
        filename
    )
    if "MW" in export_df.columns:
        export_df["MW"] = export_df["MW"].apply(
            lambda x: float(Decimal(str(x)).quantize(TWOPLACES))
        )
    #export_df.to_csv(filepath.replace(".xlsx", ".csv"), index=False)
    export_df.to_excel(filepath, index=False)
    apply_german_excel_format(filepath)

    latest_run["energy_type"] = energy_type
    if energy_type == "gas":
        latest_run["filepath"] = filepath
    if energy_type == "strom" and filepath[-10] == "d": #(stündlich)
        latest_run["filepath"] = filepath
    print(f"✔️ Excel gespeichert: {filepath}")

def apply_german_excel_format(filepath):

    wb = load_workbook(filepath)

    ws = wb.active

    for col in ws.columns:

        header = str(col[0].value).strip()

        if header == "Timestamp":
            for cell in col[1:]:
                cell.number_format = "dd.mm.yyyy hh:mm"
        
        elif header == "Tag":
            for cell in col[1:]:
                cell.number_format = "@"
        
        elif header == "Monat":
            for cell in col[1:]:
                cell.number_format = "@"

        elif header == "MW":
            for cell in col[1:]:
                cell.number_format = "0.000"

    wb.save(filepath)

    
# =========================
# ROUTES
# =========================

# HOME = STROMSEITE
@app.route("/")
def home():
    return open("strom.html", encoding="utf-8").read()


# EXPLIZITE STROMSEITE
@app.route("/strom")
def strom():
    return open("strom.html", encoding="utf-8").read()


# GASSEITE
@app.route("/gas")
def gas():
    return open("gas.html", encoding="utf-8").read()


# =========================
# CHART + SUMME
# =========================
@app.route("/print-sum", methods=["POST"])
def print_sum():
    def detect_interval_hours(timestamps):
    
        if not timestamps or len(timestamps) < 2:
            return 1
    
        ts = pd.to_datetime(pd.Series(timestamps), errors="coerce").dropna()
        ts = ts.sort_values().reset_index(drop=True)
    
        if len(ts) < 2:
            return 1
    
        # 🔥 FIX: explizit TZ setzen
        ts = pd.to_datetime(
            pd.Series(timestamps),
            errors="coerce"
        ).dropna()
        ts = ts.sort_values().reset_index(drop=True)
        diffs = ts.diff().dt.total_seconds().div(3600)
    
        corrected = []
    
        for d in diffs:
            if pd.isna(d):
                continue
    
            # Spring Forward (23h Tag → 2h Lücke)
            if abs(d - 2.0) < 0.01:
                corrected.append(1.0)
    
            # Fall Back (25h Tag → duplicate hour)
            elif abs(d) < 0.01:
                corrected.append(1.0)
    
            else:
                corrected.append(d)
    
        if not corrected:
            return 1
    
        median = pd.Series(corrected).median()
    
        if pd.isna(median) or median <= 0:
            return 1
    
        return float(median)
    data = request.get_json()

    # gas oder strom
    energy_type = data.get("energy_type", "strom")
    einheit = data.get("einheit")
    input_unit = einheit
    if not einheit:
        einheit = "kWh"
    # Startzeit bestimmen
    start_hour = get_start_hour(energy_type)

    forward_factor = safe_factor(
        data.get("forward_market_share")
    )

    raw_input = data.get("raw_input", "")
    

    # =========================
    # INTERVAL DETECTION (CRITICAL)
    # =========================
    parsed = parse_lastgang(
        raw_input,
        forward_factor,
        "X"
    )    
    if parsed.get("error"):
        return jsonify({
            "status": "error",
            "validation_error": True,
            "message": parsed["message"]
        })
    interval_hours = detect_interval_hours(parsed["timestamps"])


    if int(interval_hours) < 23:
        granularity = "1h"
    if int(interval_hours) > 23:
        granularity = "1D"
    if int(interval_hours) > 700:
        granularity = "1M"
    
    parsed = parse_lastgang(
        raw_input,
        forward_factor,
        granularity
    )

    if parsed.get("error"):
        return jsonify({
            "status": "error",
            "validation_error": True,
            "message": parsed["message"]
        })
    
    if not interval_hours or interval_hours <= 0:
        interval_hours = 1
    
    # Schutz gegen kaputte Daten (z.B. 15min / 30min / Mix)
    if interval_hours not in [1, 2, 3, 4, 6, 12, 24]:
        interval_hours = 1
    
    # HARD FIX: realistische Begrenzung
    if not interval_hours or interval_hours <= 0:
        interval_hours = 1
    
    # typische Profile absichern
    if interval_hours > 6:
        interval_hours = 1
        
    latest_run["einheit"] = "MW"
        
    lastgang_source = convert_to_mw(
        parsed["values"],
        input_unit,
        granularity,
        energy_type = energy_type,
        timestamps=parsed["timestamps"]
    )
    
    # NEU: Rundung für Summenlogik (Frontend Anzeige)
    #lastgang_source = round_for_sum(
    #    lastgang_source,
    #    input_unit
    #)
    file_timestamps = parsed["timestamps"]
    validation_result = None
    
    if (
        granularity == "1h"
        and file_timestamps
    ):
        timestamps = pd.to_datetime(
            file_timestamps,
            errors="coerce"
        ).dropna()
    
        if len(timestamps) >= 2:
    
            diffs = (
                pd.Series(timestamps)
                .sort_values()
                .diff()
                .dt.total_seconds()
                .div(3600)
                .dropna()
            )
    
            is_hourly = diffs.between(0.0, 2.01).all()
    
            if is_hourly:
                validation_result = validate_hourly_series(
                    file_timestamps
                )

    file_year_start = None
    file_year_end = None
    first_timestamp = None
    last_timestamp = None

    if validation_result and not validation_result["ok"]:
    
        return jsonify({
            "status": "error",
            "validation_error": True,
            "message":
                f"Ungültiger Abstand zwischen "
                f"{validation_result['prev']} und "
                f"{validation_result['curr']} "
                f"({validation_result['diff_hours']}h)"
        })
    
    if file_timestamps and len(file_timestamps) > 0:
    
        timestamps = pd.to_datetime(
            file_timestamps,
            dayfirst=True,
            errors="coerce"
        )
    
        timestamps = timestamps.dropna()
    
        if not timestamps.empty:
    
            first_timestamp = timestamps.min()
            
            last_timestamp = (
                timestamps.max()
                if granularity == "1h"
                else pd.Timestamp(timestamps.max())
                        .to_period("M")
                        .to_timestamp("M")
                        + pd.Timedelta(hours=23)
            )
            file_year_start = int(first_timestamp.year)
            file_year_end = int(last_timestamp.year)

    years = data.get("years") or []

    selected_year_start = None
    selected_year_end = None
    
    if years:
        selected_years = sorted([int(y) for y in years])
    
        selected_year_start = selected_years[0]
        selected_year_end = selected_years[-1]
        
    year_start = (
        int(years[0])
        if len(years) > 0
        else pd.Timestamp.now().year + 1
    )

    if file_timestamps:
        source_df = pd.DataFrame({
            "Timestamp": pd.to_datetime(file_timestamps),
            "MW": lastgang_source
        })

        if granularity in ["1D", "1M"]:
            source_df["Timestamp"] = pd.to_datetime(
                source_df["Timestamp"]
            )
        
            if source_df["Timestamp"].dt.tz is None:
                source_df["Timestamp"] = (
                    source_df["Timestamp"]
                    .dt.tz_localize(
                        "Europe/Berlin",
                        ambiguous="infer",
                        nonexistent="shift_forward"
                    )
                )
            if granularity == "1D" and energy_type == "gas":
                source_df["Timestamp"] = source_df["Timestamp"] + pd.Timedelta(hours=6)
            
        if energy_type == "strom":
            source_df = normalize_gas_day_window(
                source_df,
                energy_type
            )

        print(source_df)
        print(granularity)
        hourly_df, daily_df, monthly_df, vollversorgung_daily_df  = (
            build_all_granularities(
                source_df,
                input_unit,
                granularity,
                energy_type
            )
        )
        
    else:
    
        years_count = max(1, len(years))
        values_count = len(lastgang_source)
    
        # Granularität aus Anzahl der Werte ableiten
        if values_count == 12 * years_count:
            granularity = "1M"
            freq = "MS"
    
        elif values_count in (
            365 * years_count,
            366 * years_count
        ):
            granularity = "1D"
            freq = "D"
    
        else:
            granularity = "1h"
            freq = "h"
    
        start_ts = pd.Timestamp(
            f"{year_start}-01-01 {start_hour}:00",
            tz="Europe/Berlin"
        )
    
        timestamps = pd.date_range(
            start=start_ts,
            periods=values_count,
            freq=freq,
            tz="Europe/Berlin"
        )
    
        hourly_df = pd.DataFrame({
            "Timestamp": timestamps,
            "MW": lastgang_source
        })

    
        hourly_df, daily_df, monthly_df, vollversorgung_daily_df = (
            build_all_granularities(
                hourly_df,
                input_unit,
                granularity,
                energy_type
            )
        )
            


    return jsonify({
        "status": "ok",
    
        "granularity": granularity,   # 🔴 BOMBE
    
        "hourly_timestamps": hourly_df["Timestamp"].dt.strftime("%Y-%m-%d %H:%M").tolist() if hourly_df["Timestamp"] is not None else None,
        "hourly_values": hourly_df["MW"].tolist() if hourly_df["MW"] is not None else None,
    
        "daily_timestamps": daily_df["Timestamp"].dt.strftime("%Y-%m-%d").tolist(),
        "daily_values": daily_df["MW"].tolist(),
    
        "months": (
            monthly_df["Timestamp"]
            .dropna()
            .dt.strftime("%Y-%m")
            .tolist()
        ),
        "monthly_values": monthly_df["MW"].tolist(),
    
        "file_year_start": file_year_start,
        "file_year_end": file_year_end,
        "selected_year_start": selected_year_start,
        "selected_year_end": selected_year_end,
    
        "first_timestamp": first_timestamp.strftime("%Y-%m-%d %H:%M") if first_timestamp is not None else None,
        "last_timestamp": last_timestamp.strftime("%Y-%m-%d %H:%M") if last_timestamp is not None else None,
        
        "vollversorgung_daily_timestamps":
            vollversorgung_daily_df["Timestamp"]
                .dt.strftime("%Y-%m-%d")
                .tolist()
            if vollversorgung_daily_df is not None else None,
        
        "vollversorgung_daily_values":
            vollversorgung_daily_df["MW"].tolist()
            if vollversorgung_daily_df is not None else None,
    })


# =========================
# EXCEL EXPORT ENDPOINT
# =========================
@app.route("/save-excel", methods=["POST"])
def save_excel():
    data = request.get_json()
    selection_mode = data.get(
        "selection_mode",
        "fp_spot"
    )
    print("SELECTION MODE =", selection_mode)

    
    def detect_interval_hours(timestamps):
    
        if not timestamps or len(timestamps) < 2:
            return 1
    
        ts = pd.to_datetime(pd.Series(timestamps), errors="coerce").dropna()
        ts = ts.sort_values().reset_index(drop=True)
    
        if len(ts) < 2:
            return 1
    
        # 🔥 FIX: explizit TZ setzen
        ts = pd.to_datetime(
            pd.Series(timestamps),
            errors="coerce"
        ).dropna()
        
        ts = ts.sort_values().reset_index(drop=True)
        
        diffs = ts.diff().dt.total_seconds().div(3600)
        if ts.empty:
            return 1
    
        diffs = ts.diff().dt.total_seconds().div(3600)
    
        corrected = []
    
        for d in diffs:
            if pd.isna(d):
                continue
    
            if abs(d - 2.0) < 0.01:
                corrected.append(2.0)
            
            elif abs(d - 0.0) < 0.01:
                corrected.append(1.0)
    
            else:
                corrected.append(d)
    
        if not corrected:
            return 1
    
        median = pd.Series(corrected).median()
    
        if pd.isna(median) or median <= 0:
            return 1
    
        return float(median)
        
    # gas oder strom
    energy_type = data.get("energy_type", "strom")
    # Startzeit bestimmen
    start_hour = get_start_hour(energy_type)

    forward_factor = safe_factor(
        data.get("forward_market_share")
    )

    raw_input = data.get("raw_input", "")
    
    parsed = parse_lastgang(
        raw_input,
        forward_factor,
        "X"
    )
    
    input_unit = data.get("einheit")
    
    interval_hours = detect_interval_hours(
        parsed["timestamps"]
    )
    if interval_hours <= 0:
        interval_hours = 1

    if int(interval_hours) < 23:
        granularity = "1h"
    if int(interval_hours) > 23:
        granularity = "1D"
    if int(interval_hours) > 700:
        granularity = "1M"
    
    if energy_type == "gas" and int(interval_hours) > 23:
        parsed["timestamps"] = pd.to_datetime(parsed["timestamps"]) + pd.Timedelta(hours=5)
    
        parsed["timestamps"] = parsed["timestamps"].tolist()
    lastgang_source = convert_to_mw(
        parsed["values"],
        input_unit,
        granularity,
        energy_type=energy_type,
        timestamps=parsed["timestamps"]
    )
    # SAFETY CHECK gegen Explosion
    total = sum(lastgang_source)
    
    if total > Decimal("1_000_000_000"):
        raise Exception("Unrealistische Energiemenge erkannt – Einheit/Intervall falsch")
    
    file_timestamps = parsed["timestamps"]
    years = data.get("years") or []
    
    if not isinstance(years, list):
        years = []
        
    year_start = (
        int(years[0])
        if len(years) > 0
        else pd.Timestamp.now().year + 1
    )

    if file_timestamps:
        lastgang_source = pd.DataFrame({
            "Timestamp": pd.to_datetime(file_timestamps),
            "MW": lastgang_source
        })
        hourly_df, daily_df, monthly_df, vollversorgung_daily_df  = (
            build_all_granularities(
                lastgang_source,
                input_unit,
                granularity,
                energy_type
            )
        )

        


        
# KEIN TZ-REPROCESSING in save-excel

    else:
    
        years_count = max(1, len(years))
        values_count = len(lastgang_source)
    
        if values_count == 12 * years_count:
            granularity = "1M"
            freq = "MS"
    
        elif values_count in (
            365 * years_count,
            366 * years_count
        ):
            granularity = "1D"
            freq = "D"
    
        else:
            granularity = "1h"
            freq = "h"
    
        start_ts = pd.Timestamp(
            f"{year_start}-01-01 {start_hour}:00",
            tz=ZoneInfo("Europe/Berlin")
        )
    
        timestamps = pd.date_range(
            start=start_ts,
            periods=values_count,
            freq=freq
        )
    
        source_df = pd.DataFrame({
            "Timestamp": timestamps,
            "MW": lastgang_source
        })
    
        hourly_df, daily_df, monthly_df, vollversorgung_daily_df = (
            build_all_granularities(
                source_df,
                input_unit,
                granularity,
                energy_type
            )
        )
    
    if "Timestamp" in daily_df.columns:
        daily_df = daily_df.rename(columns={"Timestamp": "Tag"})

    if "Timestamp" in daily_df.columns:
        monthly_df = monthly_df.rename(columns={"Timestamp": "Monat"})
        
    if "Hours" in daily_df.columns:
        daily_df = daily_df.drop(columns=["Hours"])
        
    daily_df["Tag"] = pd.to_datetime(daily_df["Tag"]).dt.strftime("%d.%m.%Y")
    monthly_df = monthly_df.rename(columns={"Timestamp": "Monat"})
    monthly_df["Monat"] = monthly_df["Monat"].dt.strftime("%m.%Y")
    pricing_name = data.get("pricing_name") or "Bepreisung"
    
    timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M")
    
    energy_folder = "Gas" if energy_type == "gas" else "Strom"
    
    export_dir = os.path.join(
        os.getcwd(),
        "pricing-files",
        energy_folder,
        f"{pricing_name}_{timestamp}"
    )
    
    #os.makedirs(export_dir, exist_ok=True)
    
    # =========================
    # FIX: EXCEL ERZEUGUNG IN DEFINIERTER REIHENFOLGE
    # =========================
    
    # 1) zuerst: Original-Granularität (stündlich = Basis)
    hourly_file = save_excel_file(
        hourly_df,
        pricing_name,
        energy_type,
        "stuendlich",
        export_dir
    )
    
    # 2) dann: aggregiert aus Basis (nur wenn feiner → erlaubt)
    daily_file = save_excel_file(
        daily_df,
        pricing_name,
        energy_type,
        "taeglich",
        export_dir
    )
    
    # 3) zuletzt: weitere Aggregation
    monthly_file = save_excel_file(
        monthly_df,
        pricing_name,
        energy_type,
        "monatlich",
        export_dir
    )
        

    
    files = sorted(
        [
            os.path.join(export_dir, f)
            for f in os.listdir(export_dir)
            if pricing_name in f
        ],
        key=os.path.getmtime,
        reverse=True
    )
    
    hourly_file = next(
        (f for f in files if "_stuendlich.xlsx" in f),
        None
    )
    
    daily_file = next(
        (f for f in files if "_taeglich.xlsx" in f),
        None
    )
    
    monthly_file = next(
        (f for f in files if "_monatlich.xlsx" in f),
        None
    )

    return jsonify({
        "status": "ok",
        "files": [
            f"{energy_type}_{pricing_name}_stuendlich.xlsx",
            f"{energy_type}_{pricing_name}_taeglich.xlsx",
            f"{energy_type}_{pricing_name}_monatlich.xlsx"
        ]
    })

pricing_status = {
    "running": False,
    "completed": False,
    "error": None,

    "step": 0,
    "message": "",

    "result": None
}

# =========================
# PRICING SIMULATION
# =========================



SELENIUM_URL = "seleniumurl"

GAS_PRICING_URL = "gasurl"
STROM_PRICING_URL = "stromurl"

USERNAME = "mailaddr"
PASSWORD = "passwd"

OUTPUTS = []
def run_pricing():

    file_path = latest_run.get("filepath")
    einheit = latest_run.get("einheit")
    energy_type = latest_run.get("energy_type")
    print("▶ Selenium Run Context:")
    print("file_path:", file_path)
    print("einheit:", einheit)
    print("energy_type:", energy_type)
    if not file_path:
        raise Exception("Kein Datei-Pfad verfügbar (latest_run leer)")

    options = webdriver.ChromeOptions()

    options.add_argument("--no-sandbox")

    driver = webdriver.Remote(
        command_executor=SELENIUM_URL,
        options=options
    )

    try:

        pricing_status["message"] = "Marktdaten werden geladen und Monatsprofil vorbereitet ..."

        filename = os.path.basename(file_path)
        
        if filename.startswith("Gas"):
            driver.get(GAS_PRICING_URL)
        elif filename.startswith("Strom"):
            driver.get(STROM_PRICING_URL)
        else:
            raise Exception(f"Unbekannter Dateityp: {filename}")

        driver.find_element(
            By.ID,
            "username"
        ).send_keys(USERNAME)

        driver.find_element(
            By.ID,
            "password"
        ).send_keys(PASSWORD)

        driver.find_element(
            By.ID,
            "kc-login"
        ).click()

        pricing_status["message"] = "Forwardkurven werden berechnet"

        body = driver.find_element(
            By.TAG_NAME,
            "body"
        )

        driver.find_element(
            By.ID,
            "Offer_Indicative_Binding_1"
        ).click()
        #driver.save_screenshot("l1.png")  #############################################
        unit = "MW"
        
        # Dropdown öffnen (das aktuell MW anzeigt)
        driver.find_element(
            By.XPATH,
            "//*[contains(normalize-space(.), 'MW')]"
        ).click()
        
        time.sleep(1)   
        # Jetzt nach der gewünschten Option suchen
        target = driver.find_element(
            By.XPATH,
            f"//*[normalize-space(text())='{unit}']"
        )
        
        driver.execute_script(
            "arguments[0].click();",
            target
        )
        
        if target:
            driver.execute_script("arguments[0].click();", target)
        else:
            print(f"⚠️ Einheit {unit} nicht gefunden – kein Fallback mehr auf kWh!")
        #driver.save_screenshot("l2.png")  #############################################
        unit = "MW"
        
        elements = driver.find_elements(By.XPATH, "//*")
        
        target = None
        
        for e in elements:
            try:
                text = (e.text or "").strip()
        
                if unit in text:
                    print("FOUND:", repr(text))
                    target = e
                    break
        
            except Exception:
                pass
        
        if target:
            driver.execute_script(
                "arguments[0].scrollIntoView(true);",
                target
            )
            time.sleep(0.5)
        
            driver.execute_script(
                "arguments[0].click();",
                target
            )
        
            print("✔ Einheit gewählt:", unit)
        
        else:
            print("❌ Einheit nicht gefunden:", unit)
        #driver.save_screenshot("l3.png")  #############################################
        file_input = driver.find_element(
            By.ID,
            "ProfileRawFile"
        )

        file_input.send_keys(
            os.path.abspath(file_path)
        )

        body.send_keys(Keys.ENTER)

        time.sleep(1)      
        #driver.save_screenshot("l4.png")  #############################################
        driver.execute_script("""
        let el = document.querySelector(
            'a[onclick="OnBootstrapDropdownItemClick(this)"]'
        );
        if(el) el.click();
        """)
        #driver.save_screenshot("l5.png")  #############################################
        time.sleep(1)    
        pricing_status["message"] = "Strukturierungsaufschlag ermittelt"

        time.sleep(1)    
        #driver.save_screenshot("l6.png")  #############################################
        driver.find_element(
            By.ID,
            "btnSubmit"
        ).click()
        
        if filename.startswith("Gas"):
            time.sleep(2)
        elif filename.startswith("Strom"):
            time.sleep(40)
        #driver.save_screenshot("l7.png")  #############################################
        now = pd.Timestamp.now() + pd.Timedelta(hours=4)
        
        validity = now.ceil("30min")
        
        # Preis aus Selenium auslesen
        current_price = Decimal(
            driver.find_element(By.ID, "Price").get_attribute("value")
        )
        
        result = {
            "price": str(current_price).replace(".", ","),
            "amount": driver.find_element(By.ID, "Amount").get_attribute("value"),
            "validity": validity.strftime("%d%m%Y %H%M"), # "—",
            "psr": driver.find_element(By.ID, "PSR").get_attribute("value")
        }
        
        # Prüfen, ob bereits ein Ergebnis vorhanden ist
        price_improved = False
        previous_price = None
        
        if len(OUTPUTS) == 0:
        
            selected_result = result
        
        else:
        
            previous_price = Decimal(
                OUTPUTS[-1]["price"].replace(",", ".")
            )
        
            if current_price < previous_price:
                selected_result = result
                price_improved = True
            else:
                selected_result = OUTPUTS[-1]
        
        OUTPUTS.append(selected_result)
        
        return {
            "result": selected_result,
            "price_improved": price_improved,
            "previous_price": (
                str(previous_price).replace(".", ",")
                if previous_price is not None
                else None
            )
        }
        print("Aktueller Preis:", current_price)
        print("Bisher beste Ergebnisse:", OUTPUTS)
        

    
    finally:
        driver.quit()




def run_pricing_background():
    global pricing_status, OUTPUTS
    pricing_status["price_improved"] = False
    try:
        pricing_status["running"] = True
        pricing_status["completed"] = False
        pricing_status["error"] = None
        pricing_status["step"] = 0
        pricing_status["result"] = None
        pricing_status["message"] = "Pricing wird vorbereitet"
    
        # Ergebnisse eines neuen Pricing-Durchlaufs zurücksetzen
        OUTPUTS.clear()
    
        # Insgesamt 5 Pricing-Läufe
        for i in range(5):
    
            pricing_status["step"] = i + 1
            pricing_status["message"] = (
                f"Pricing-Lauf {i + 1} von 5 wird ausgeführt"
            )
    
            print(
                f"========== PRICING LAUF {i + 1} VON 5 =========="
            )
    
            # run_pricing() bleibt komplett unverändert
            run_result = run_pricing()
            # WICHTIG:
            # Ergebnis sofort nach diesem Lauf speichern.
            # Das Frontend holt es über /pricing/status
            pricing_status["result"] = run_result["result"]
            pricing_status["price_improved"] = run_result["price_improved"]
            pricing_status["previous_price"] = run_result["previous_price"]

    
            pricing_status["message"] = (
                f"Pricing-Lauf {i + 1} von 5 abgeschlossen"
            )
    
            print(
                f"========== PRICING LAUF {i + 1} ABGESCHLOSSEN =========="
            )
            print("Ergebnis:", pricing_status["result"])
    
        # Alle 5 Läufe abgeschlossen
        pricing_status["completed"] = True
        pricing_status["running"] = False
        pricing_status["step"] = 5
        pricing_status["message"] = "Alle 5 Pricing-Läufe abgeschlossen"
    
    except Exception as e:
    
        pricing_status["running"] = False
        pricing_status["completed"] = False
        pricing_status["error"] = str(e)
        pricing_status["message"] = str(e)
    
        print("❌ Pricing Fehler:", e)
    
    





        
@app.route("/pricing/start", methods=["POST"])
def pricing_start():

    global pricing_status

    if pricing_status["running"]:

        return jsonify({
            "status": "already_running"
        })

    executor = ThreadPoolExecutor(max_workers=1)
    import time
    latest_run["run_id"] = str(time.time())
    executor.submit(
        run_pricing_background
    )

    return jsonify({
        "status": "started"
    })


@app.route("/pricing/status", methods=["GET"])
def pricing_status_route():

    global pricing_status

    return jsonify(pricing_status)

@app.route("/pricing/result", methods=["GET"])
def pricing_result():

    global pricing_status

    if not pricing_status["completed"]:

        return jsonify({
            "status": "not_finished"
        })

    return jsonify({
        "status": "completed",
        "result": pricing_status["result"]
    })


# =========================
# VV-AUFSCHLAG AUS EXCEL
# =========================

@app.route("/aufschlaege", methods=["GET"])
@app.route("/aufschlaege", methods=["GET"])
def aufschlaege():

    filepath = "excel.xlsm"

    print("===================================")
    print("AUFSCHLÄGE ROUTE AUFGERUFEN")
    print("Pfad:", filepath)
    print("Existiert:", os.path.isfile(filepath))

    if not os.path.isfile(filepath):
        return jsonify({
            "status": "error",
            "message": "Excel-Datei nicht gefunden",
            "path": filepath
        }), 404

    try:
        wb = load_workbook(
            filepath,
            read_only=True,
            data_only=True,
            keep_vba=True
        )

        ws = wb["Ausgabe EnS"]
        value26 = str(round(ws["C7"].value, 2)).replace(".", ",")
        value27 = str(round(ws["C16"].value, 2)).replace(".", ",")
        value28 = str(round(ws["C25"].value, 2)).replace(".", ",")
        value29 = str(round(ws["C34"].value, 2)).replace(".", ",")

        bf26 = str(round(ws["C8"].value, 2)).replace(".", ",")
        bf27 = str(round(ws["C17"].value, 2)).replace(".", ",")
        bf28 = str(round(ws["C26"].value, 2)).replace(".", ",")
        bf29 = str(round(ws["C35"].value, 2)).replace(".", ",")
        
        data = request.args
        years = data.getlist("years")

        result_vv = []
        result_bf = []
        if len(years) == 0:
            year = "Jahr"
            value = "fehlt"
        else:
            for year in years:
                year = int(year)
    
                if year == 2026:
                    vv_value = value26
                    bf_value = bf26
                elif year == 2027:
                    vv_value = value27
                    bf_value = bf27
                elif year == 2028:
                    vv_value = value28
                    bf_value = bf28
                elif year == 2029:
                    vv_value = value29
                    bf_value = bf29
                else:
                    vv_value = "-"
                    bf_value = "-"
    
                result_vv.append(f"<small>{year}</small>: {vv_value}")
                result_bf.append(f"<small>{year}</small>: {bf_value}")

        wb.close()


        return jsonify({
            "status": "ok",
            "aufschlaege": "\n".join(result_vv),
            "bf_aufschlaege": "\n".join(result_bf)
        }), 200

    except Exception as e:
        print("AUFSCHLAG FEHLER:", repr(e))

        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500
# =========================
# START
# =========================
# Checking if an argument is provided before accessing it   
if len(sys.argv) > 1:   
  portno = int(sys.argv[1])
else:
  portno = 27274
if __name__ == "__main__":
    
    print("===================================")
    print("REGISTRIERTE FLASK ROUTEN:")
    print(app.url_map)
    print("===================================")
        
    app.run(
        host="0.0.0.0",
        port=portno,
        debug=False,
        use_reloader=False
    )
